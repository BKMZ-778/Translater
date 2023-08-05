import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

def start_ozon():
    fileName = filedialog.askopenfilename()
    df_ozon_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L,M,N', skiprows=1)
    df_ozon_file = df_ozon_file.rename(columns={11: 'bad_description', 12: 'price', 13: 'link'})
    print(df_ozon_file)
    df_baza_ozon = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx', sheet_name=0, engine='openpyxl', header=None, usecols='A,B', skiprows=1)
    df_baza_ozon = df_baza_ozon.rename(columns={0: 'bad_description', 1: 'good_description'})
    print(df_baza_ozon)
    df_merged = pd.merge(df_ozon_file, df_baza_ozon, how='left', left_on='bad_description', right_on='bad_description')

    df_to_translate = df_merged.loc[df_merged['good_description'].isnull()]
    df_to_translate['good_description'] = df_to_translate['bad_description']
    df_to_translate = df_to_translate[['bad_description', 'good_description', 'price', 'link']]
    df_to_translate = df_to_translate.drop_duplicates(subset='good_description', keep='first').sort_values(by='good_description')
    print(df_to_translate)
    writer = pd.ExcelWriter(f'{fileName}-На перевод.xlsx', engine='openpyxl')
    df_to_translate.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    wb = openpyxl.load_workbook(f'{fileName}-На перевод.xlsx')
    ws = wb.active
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9

    wb.save(f'{fileName}-На перевод.xlsx')
    msg = "Готово!"
    mb.showinfo("Информация", msg)

def add_tobaza():
    fileName = filedialog.askopenfilename()
    df_to_append = pd.read_excel(fileName)
    df_base = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx')
    df_base_updated = pd.concat([df_base, df_to_append], axis=0).drop_duplicates(subset='bad_description')
    writer = pd.ExcelWriter('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx', engine='openpyxl')
    df_base_updated.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    msg = "Добавлено!"
    mb.showinfo("Информация", msg)

    fileName = filedialog.askopenfilename()
    df_base_updated = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx')
    df_ozon_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L', skiprows=1)
    #df_ozon_file = df_ozon_file[df_ozon_file[11].notna()]
    print(df_ozon_file)
    df_ozon_file = df_ozon_file.rename(columns={11: 'bad_description'})
    df_merged = pd.merge(df_ozon_file, df_base_updated, how='left', left_on='bad_description', right_on='bad_description')
    print(df_merged)
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    ws.insert_cols(13)
    i = 2
    for row in df_merged['good_description']:
        ws[f"M{i}"].value = row
        i += 1

    wb.save(fileName)

    msg = "Обновленно!"
    mb.showinfo("Информация", msg)

def start_LD():
    fileName = filedialog.askopenfilename()
    df_LD_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L,M,N,Y', skiprows=1)
    df_LD_file = df_LD_file.rename(columns={11: 'china_description', 12: 'price', 13: 'link', 24: 'SKU'})
    df_baza_LD = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ЛД.xlsx', sheet_name=0, engine='openpyxl', header=None, usecols='A,C', skiprows=1)
    df_baza_LD = df_baza_LD.rename(columns={0: 'SKU', 2: 'good_description'})
    df_merged = pd.merge(df_LD_file, df_baza_LD, how='left', left_on='SKU', right_on='SKU')
    print(df_merged)
    df_to_translate = df_merged.loc[df_merged['good_description'].isnull()]
    print(df_to_translate)
    df_to_translate = df_to_translate[['china_description', 'good_description', 'price', 'link', 'SKU']]
    df_to_translate = df_to_translate.drop_duplicates(subset='SKU', keep='first').sort_values(by='china_description')
    print(df_to_translate)

    writer = pd.ExcelWriter(f'{fileName}-На перевод.xlsx', engine='openpyxl')
    df_to_translate.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    wb = openpyxl.load_workbook(f'{fileName}-На перевод.xlsx')
    ws = wb.active

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9

    len_sheet = ws.max_row
    ws.move_range(f"D1:D{len_sheet}", cols=10)
    ws.move_range(f"E1:E{len_sheet}", cols=7)
    wb.save(f'{fileName}-На перевод.xlsx')
    msg = "Готово!"
    mb.showinfo("Информация", msg)

window = tk.Tk()
window.title('OZON')
window.geometry("400x250+500+300")

button = tk.Button(text="Реестр OZON", width=24, height=2, bg="lightgrey", fg="black", command=start_ozon)
button.configure(font=('hank', 10))

button2 = tk.Button(text="Добавить в базу + обновить реестр", width=35, height=2, bg="lightgrey", fg="black", command=add_tobaza)
button2.configure(font=('hank', 10))

button3 = tk.Button(text="На перевод LD", width=24, height=2, bg="lightgrey", fg="black", command=start_LD)
button3.configure(font=('hank', 10))

button.pack()
button2.pack()
button3.pack()

window.mainloop()